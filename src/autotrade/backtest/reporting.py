"""Backtest history recording and post-run performance analysis."""

from __future__ import annotations

from copy import deepcopy
from dataclasses import fields, is_dataclass
from enum import Enum
import math
from pathlib import Path
import sys
from typing import TYPE_CHECKING
import pandas as pd

from autotrade.engine.security_manager import SecurityManager
from autotrade.engine.oms import OmsBase

if TYPE_CHECKING:
    from autotrade.backtest.gateway import AccountLedger


class BacktestRecorder:
    """Copy authoritative OMS/Security state without calculating it."""

    def __init__(self) -> None:
        self.account_daily: dict = {}
        self.position_daily: dict = {}
        self.contract_daily: dict = {}

    def snapshot(
        self,
        when,
        ledger: "AccountLedger",
        security_manager: SecurityManager,
    ) -> None:
        account = ledger.account
        self.account_daily[when] = {
            "cash": account.cash,
            "margin": account.margin,
            "realized_pnl": account.realized_pnl,
            "unrealized_pnl": account.unrealized_pnl,
            "equity": account.equity,
            "available": account.available,
        }
        self.position_daily[when] = deepcopy(ledger.get_all_positions())

        symbols = (
            set(ledger.positions)
            | set(ledger.realized_pnl_by_symbol)
            | set(ledger.turnover_by_symbol)
        )
        self.contract_daily[when] = {
            instrument_id: self._symbol_snapshot(instrument_id, ledger, security_manager)
            for instrument_id in sorted(symbols)
        }

    @staticmethod
    def _symbol_snapshot(
        instrument_id: str,
        ledger: "AccountLedger",
        security_manager: SecurityManager,
    ) -> dict:
        position = ledger.positions.get(instrument_id)
        security = security_manager.get(instrument_id)
        return {
            "volume": 0.0 if position is None else position.volume,
            "margin": 0.0 if position is None else position.margin,
            "realized_pnl": ledger.realized_pnl_by_symbol.get(instrument_id, 0.0),
            "unrealized_pnl": ledger.unrealized_pnl_by_symbol.get(instrument_id, 0.0),
            "turnover": ledger.turnover_by_symbol.get(instrument_id, 0.0),
            "commission": ledger.commission_by_symbol.get(instrument_id, 0.0),
            "price": None if security is None else security.price,
        }

    @staticmethod
    def get_trade_log_df(oms: OmsBase) -> pd.DataFrame:
        return pd.DataFrame([
            {
                "datetime": trade.datetime,
                "instrument_id": trade.instrument_id,
                "orderid": trade.orderid,
                "direction": trade.direction,
                "price": trade.price,
                "traded": trade.traded,
                "volume": trade.volume,
                "avgFillPrice": trade.avgFillPrice,
                "status": trade.status,
                "reference": trade.reference,
            }
            for trade in oms.get_all_trades()
        ])

    def get_account_daily_df(self) -> pd.DataFrame:
        frame = pd.DataFrame.from_dict(self.account_daily, orient="index").sort_index()
        frame.index.name = "date"
        return frame

    def get_position_daily_df(self) -> pd.DataFrame:
        """Return the append-only position snapshots as ``(date, instrument_id)`` rows."""
        records = []
        for when, positions in self.position_daily.items():
            for position in positions:
                records.append({
                    "date": when,
                    **self._record_fields(position),
                })
        if not records:
            return pd.DataFrame(index=pd.MultiIndex.from_arrays([[], []], names=("date", "instrument_id")))
        return pd.DataFrame.from_records(records).set_index(
            ["date", "instrument_id"], drop=True,
        ).sort_index()

    @staticmethod
    def _record_fields(record) -> dict:
        if is_dataclass(record):
            return {field.name: getattr(record, field.name) for field in fields(record)}
        if isinstance(record, dict):
            return dict(record)
        return dict(vars(record))


class PerformanceAnalyzer:
    """Pure post-run calculations over recorded account equity."""

    def __init__(
        self,
        *,
        initial_cash: float,
        risk_free: float = 0.02,
        annual_days: int = 252,
    ) -> None:
        self.initial_cash = float(initial_cash)
        self.risk_free = float(risk_free)
        self.annual_days = int(annual_days)
        if self.annual_days <= 0:
            raise ValueError("annual_days must be positive")

    def calculate(self, account_history: dict, *, print_result: bool = True) -> dict:
        df = pd.DataFrame.from_dict(account_history, orient="index").sort_index()
        if df.empty:
            return {}
        equity = df["equity"].astype(float)
        final_equity = float(equity.iloc[-1])
        total_return = (
            final_equity / self.initial_cash - 1
            if self.initial_cash > 0
            else math.nan
        )
        max_drawdown = self.max_drawdown(equity)
        timestamps = pd.to_datetime(equity.index)
        timed_equity = pd.Series(equity.to_numpy(), index=timestamps).sort_index()
        # Each observed calendar date represents one trading/session return.
        # ``dropna`` intentionally excludes weekends and holidays rather than
        # treating them as zero-return sessions.
        daily_equity = timed_equity.resample("1D").last().dropna()
        daily_returns = daily_equity.pct_change().dropna()
        annual_return = self.calculate_annual_return(
            self.initial_cash,
            final_equity,
            trading_periods=len(daily_returns),
            annual_days=self.annual_days,
        )
        if len(daily_returns) >= 2 and daily_returns.std() > 0:
            risk_free_period = math.expm1(math.log1p(self.risk_free) / self.annual_days)
            sharpe = (
                (daily_returns.mean() - risk_free_period)
                / daily_returns.std()
                * math.sqrt(self.annual_days)
            )
        else:
            sharpe = math.nan
        if print_result:
            print("\n===== 回测绩效 =====")
            print(f"初始资金: {self.initial_cash:.2f}")
            print(f"结束资金: {final_equity:.2f}")
            print(f"总收益率: {total_return * 100:.2f}%")
            print(f"年化收益率: {annual_return * 100:.2f}%")
            print(f"最大回撤: {max_drawdown * 100:.2f}%")
            print(f"Sharpe Ratio: {sharpe:.2f}")
        return {
            "total_return": f"{total_return * 100:.2f}%",
            "annual_return": f"{annual_return * 100:.2f}%",
            "sharpe": sharpe,
            "max_drawdown": f"{max_drawdown * 100:.2f}%",
        }

    @staticmethod
    def calculate_annual_return(
        initial_equity: float,
        final_equity: float,
        *,
        trading_periods: int,
        annual_days: int = 252,
    ) -> float:
        """Annualize return from observed trading/session return periods."""
        if (
            trading_periods <= 0
            or annual_days <= 0
            or initial_equity <= 0
            or final_equity <= 0
            or not math.isfinite(initial_equity)
            or not math.isfinite(final_equity)
        ):
            return math.nan
        years = trading_periods / annual_days
        annual_log_return = (
            math.log(final_equity) - math.log(initial_equity)
        ) / years
        if annual_log_return > math.log(sys.float_info.max):
            return math.inf
        return math.expm1(annual_log_return)

    @staticmethod
    def max_drawdown(equity_series) -> float:
        peak = float(equity_series.iloc[0])
        maximum = 0.0
        for equity in equity_series:
            peak = max(peak, float(equity))
            maximum = max(maximum, (peak - float(equity)) / peak)
        return maximum


class BacktestReporting:
    """Facade for backtest history export and post-run statistics."""

    def __init__(
        self,
        *,
        recorder: BacktestRecorder,
        analyzer: PerformanceAnalyzer,
        oms: OmsBase,
    ) -> None:
        self.recorder = recorder
        self.analyzer = analyzer
        self.oms = oms
        self.result: dict = {}

    def calculate(self, *, print_result: bool = True) -> dict:
        if not self.recorder.account_daily:
            self.result = {}
        else:
            self.result = self.analyzer.calculate(
                self.recorder.account_daily,
                print_result=print_result,
            )
        return self.result

    def get_trade_log_df(self) -> pd.DataFrame:
        return self.recorder.get_trade_log_df(self.oms)

    def get_account_daily_df(self) -> pd.DataFrame:
        return self.recorder.get_account_daily_df()

    def get_position_daily_df(self) -> pd.DataFrame:
        return self.recorder.get_position_daily_df()

    def export_xlsx(self, path: str | Path) -> Path:
        """Write the base backtest report to a four-sheet Excel workbook.

        ``path`` is an explicit output location and its parent directories are
        created when necessary.  Subclasses extend ``_export_frames`` to add
        domain-specific sheets without duplicating workbook logic.
        """
        target = Path(path)
        if target.suffix.lower() != ".xlsx":
            raise ValueError("report export path must end with .xlsx")
        target.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            for sheet_name, frame in self._export_frames().items():
                safe = self._excel_safe_frame(frame)
                # Repeated MultiIndex labels are facts, not visual grouping:
                # write every timestamp explicitly so a round-trip through
                # Excel does not turn rows into NaT when read back.
                safe.to_excel(
                    writer, sheet_name=sheet_name, index=True, merge_cells=False,
                )
            self._format_export_workbook(writer)
        return target

    def _export_frames(self) -> dict[str, pd.DataFrame]:
        result = self.calculate(print_result=False)
        return {
            "performance": pd.DataFrame.from_dict(result, orient="index", columns=["value"]),
            "account_daily": self.get_account_daily_df(),
            "trade_log": self.get_trade_log_df(),
            "position_daily": self.get_position_daily_df(),
        }

    @staticmethod
    def _excel_safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
        """Convert non-Excel values such as enums and missing-reason tuples."""
        def excel_value(value):
            if isinstance(value, Enum):
                return value.value
            if isinstance(value, (tuple, list, set)):
                return "; ".join(map(str, value))
            if isinstance(value, dict):
                return str(value)
            return value

        return frame.map(excel_value)

    @staticmethod
    def _format_export_workbook(writer) -> None:
        from openpyxl.styles import Font, PatternFill

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
            for column_cells in worksheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 40)
                worksheet.column_dimensions[column_cells[0].column_letter].width = width


__all__ = [
    "BacktestRecorder",
    "BacktestReporting",
    "PerformanceAnalyzer",
]
